"""
Excel Sync Service for Insurance Portal
Integrates with Google Drive for shared Excel file access
"""

import os
import time
import threading
import logging
from datetime import datetime
from pathlib import Path
from supabase import create_client, Client
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import hashlib
import json
from config import Config
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io

logger = logging.getLogger(__name__)

class ExcelSyncService:
    def __init__(self, excel_filename="insurance_data.xlsx"):
        """Initialize Excel sync service with Google Drive integration"""
        self.supabase: Client = create_client(Config.SUPABASE_URL, Config.SUPABASE_KEY)
        self.excel_filename = excel_filename
        self.local_excel_path = os.path.join(os.getcwd(), excel_filename)
        self.sync_lock = threading.Lock()
        
        # Initialize Google Drive service
        self.drive_service = self._init_google_drive()
        self.excel_file_id = None
        
        # Find or create Excel file in Google Drive root
        self._setup_drive_excel_file()
        
        logger.info(f"Excel sync service initialized for file: {excel_filename}")

    def _init_google_drive(self):
        """Initialize Google Drive API service"""
        try:
            credentials = Credentials.from_service_account_file(
                Config.GOOGLE_CREDENTIALS_FILE,
                scopes=['https://www.googleapis.com/auth/drive']
            )
            service = build('drive', 'v3', credentials=credentials)
            logger.info("Google Drive service initialized successfully")
            return service
        except Exception as e:
            logger.error(f"Failed to initialize Google Drive service: {e}")
            raise

    def _setup_drive_excel_file(self):
        """Find or create Excel file in Google Drive root"""
        try:
            # Search for existing Excel file in root
            query = f"name='{self.excel_filename}' and parents in 'root' and trashed=false"
            results = self.drive_service.files().list(q=query).execute()
            files = results.get('files', [])
            
            if files:
                self.excel_file_id = files[0]['id']
                logger.info(f"Found existing Excel file in Drive: {self.excel_file_id}")
            else:
                # Create new Excel file
                self._create_initial_excel_file()
                logger.info(f"Created new Excel file in Drive: {self.excel_file_id}")
                
        except Exception as e:
            logger.error(f"Error setting up Drive Excel file: {e}")
            raise

    def _create_initial_excel_file(self):
        """Create initial Excel file with current Supabase data"""
        try:
            # Generate Excel locally first
            self._export_supabase_to_local_excel()
            
            # Upload to Google Drive root
            file_metadata = {
                'name': self.excel_filename,
                'parents': ['root']  # Root folder
            }
            
            media = MediaFileUpload(
                self.local_excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            file = self.drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            
            self.excel_file_id = file.get('id')
            
            # Set permissions for anyone with link to edit
            permission = {
                'type': 'anyone',
                'role': 'writer'
            }
            self.drive_service.permissions().create(
                fileId=self.excel_file_id,
                body=permission
            ).execute()
            
            logger.info(f"Excel file created and shared in Google Drive root")
            
        except Exception as e:
            logger.error(f"Error creating initial Excel file: {e}")
            raise

    def _format_excel(self, filename):
        """Apply formatting to Excel file"""
        try:
            wb = load_workbook(filename)
            
            for sheet in wb.worksheets:
                # Format headers
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            logger.debug(f"Excel formatting applied to {filename}")
        except Exception as e:
            logger.warning(f"Could not format Excel: {e}")

    def _export_supabase_to_local_excel(self):
        """Export Supabase data to local Excel file"""
        try:
            logger.info("Exporting Supabase data to local Excel...")
            
            with pd.ExcelWriter(self.local_excel_path, engine='openpyxl') as writer:
                # Clients
                clients = self.supabase.table("clients").select("*").execute()
                df_clients = pd.DataFrame(clients.data)
                if not df_clients.empty:
                    df_clients.to_excel(writer, sheet_name="Clients", index=False)
                
                # Members with client info
                members = self.supabase.table("members").select(
                    "*, clients(name, phone, email)"
                ).execute()
                df_members = pd.DataFrame(members.data)
                if not df_members.empty:
                    # Process client join data
                    if 'clients' in df_members.columns:
                        client_df = pd.json_normalize(df_members['clients'])
                        client_df.columns = ['client_' + col for col in client_df.columns]
                        df_members = pd.concat([df_members.drop('clients', axis=1), client_df], axis=1)
                    
                    # Reorder columns
                    cols = df_members.columns.tolist()
                    if 'client_id' in cols:
                        cols.remove('client_id')
                        if 'member_id' in cols:
                            idx = cols.index('member_id') + 1
                            cols.insert(idx, 'client_id')
                        else:
                            cols.insert(0, 'client_id')
                        df_members = df_members[cols]
                    
                    df_members.to_excel(writer, sheet_name="Members", index=False)
                
                # Policies with client and member info
                policies = self.supabase.table("policies").select(
                    "*, clients(name, phone, email), members(member_name)"
                ).execute()
                df_policies = pd.DataFrame(policies.data)
                if not df_policies.empty:
                    # Process join data
                    if 'clients' in df_policies.columns:
                        client_df = pd.json_normalize(df_policies['clients'])
                        client_df.columns = ['client_' + col for col in client_df.columns]
                        df_policies = pd.concat([df_policies.drop('clients', axis=1), client_df], axis=1)
                    
                    if 'members' in df_policies.columns:
                        member_df = pd.json_normalize(df_policies['members'])
                        member_df.columns = ['member_' + col for col in member_df.columns]
                        df_policies = pd.concat([df_policies.drop('members', axis=1), member_df], axis=1)
                    
                    # Add commission amount calculation (Net + Addon) * Commission %
                    if 'net_premium' in df_policies.columns and 'commission_percentage' in df_policies.columns:
                        net_premium = pd.to_numeric(df_policies['net_premium'], errors='coerce').fillna(0)
                        addon_premium = pd.to_numeric(df_policies.get('addon_premium', 0), errors='coerce').fillna(0)
                        commission_percentage = pd.to_numeric(df_policies['commission_percentage'], errors='coerce').fillna(0)
                        
                        # Calculate commission based on (Net + Addon) * Commission %
                        commission_base = net_premium + addon_premium
                        df_policies['calculated_commission_amount'] = (commission_base * commission_percentage / 100).round(2)
                        
                        # Use stored commission_amount if available, otherwise use calculated
                        if 'commission_amount' not in df_policies.columns:
                            df_policies['commission_amount'] = df_policies['calculated_commission_amount']
                        else:
                            # Fill missing commission_amount with calculated values
                            df_policies['commission_amount'] = df_policies['commission_amount'].fillna(df_policies['calculated_commission_amount'])
                        
                        # Remove the temporary calculated column
                        df_policies.drop('calculated_commission_amount', axis=1, inplace=True)
                    
                    # Reorder columns
                    cols = df_policies.columns.tolist()
                    if 'client_id' in cols and 'member_id' in cols:
                        cols.remove('client_id')
                        cols.remove('member_id')
                        if 'policy_id' in cols:
                            idx = cols.index('policy_id') + 1
                            cols.insert(idx, 'client_id')
                            cols.insert(idx + 1, 'member_id')
                        else:
                            cols.insert(0, 'client_id')
                            cols.insert(1, 'member_id')
                        df_policies = df_policies[cols]
                    
                    df_policies.to_excel(writer, sheet_name="Policies", index=False)
                
                # Claims with policy and client info
                claims = self.supabase.table("claims").select(
                    "*, policies(policy_number, clients(name, phone, email))"
                ).execute()
                df_claims = pd.DataFrame(claims.data)
                if not df_claims.empty:
                    # Process join data
                    if 'policies' in df_claims.columns:
                        policy_df = pd.json_normalize(df_claims['policies'])
                        policy_df.columns = ['policy_' + col for col in policy_df.columns]
                        df_claims = pd.concat([df_claims.drop('policies', axis=1), policy_df], axis=1)
                    
                    df_claims.to_excel(writer, sheet_name="Claims", index=False)
                
                # Pending Policies
                pending = self.supabase.table("pending_policies").select(
                    "*, clients(name, phone, email), members(member_name)"
                ).execute()
                df_pending = pd.DataFrame(pending.data)
                if not df_pending.empty:
                    # Process join data
                    if 'clients' in df_pending.columns:
                        client_df = pd.json_normalize(df_pending['clients'])
                        client_df.columns = ['client_' + col for col in client_df.columns]
                        df_pending = pd.concat([df_pending.drop('clients', axis=1), client_df], axis=1)
                    
                    if 'members' in df_pending.columns:
                        member_df = pd.json_normalize(df_pending['members'])
                        member_df.columns = ['member_' + col for col in member_df.columns]
                        df_pending = pd.concat([df_pending.drop('members', axis=1), member_df], axis=1)
                    
                    # Add commission amount calculation for pending policies (Net + Addon) * Commission %
                    if 'net_premium' in df_pending.columns and 'commission_percentage' in df_pending.columns:
                        net_premium = pd.to_numeric(df_pending['net_premium'], errors='coerce').fillna(0)
                        addon_premium = pd.to_numeric(df_pending.get('addon_premium', 0), errors='coerce').fillna(0)
                        commission_percentage = pd.to_numeric(df_pending['commission_percentage'], errors='coerce').fillna(0)
                        
                        # Calculate commission based on (Net + Addon) * Commission %
                        commission_base = net_premium + addon_premium
                        df_pending['calculated_commission_amount'] = (commission_base * commission_percentage / 100).round(2)
                        
                        # Use stored commission_amount if available, otherwise use calculated
                        if 'commission_amount' not in df_pending.columns:
                            df_pending['commission_amount'] = df_pending['calculated_commission_amount']
                        else:
                            # Fill missing commission_amount with calculated values
                            df_pending['commission_amount'] = df_pending['commission_amount'].fillna(df_pending['calculated_commission_amount'])
                        
                        # Remove the temporary calculated column
                        df_pending.drop('calculated_commission_amount', axis=1, inplace=True)
                    
                    # Reorder columns
                    cols = df_pending.columns.tolist()
                    if 'client_id' in cols and 'member_id' in cols:
                        cols.remove('client_id')
                        cols.remove('member_id')
                        if 'pending_id' in cols:
                            idx = cols.index('pending_id') + 1
                            cols.insert(idx, 'client_id')
                            cols.insert(idx + 1, 'member_id')
                        else:
                            cols.insert(0, 'client_id')
                            cols.insert(1, 'member_id')
                        df_pending = df_pending[cols]
                    
                    df_pending.to_excel(writer, sheet_name="Pending Policies", index=False)
                
                # Health Insurance Details with floater fields
                health_details = self.supabase.table("health_insurance_details").select("*").execute()
                df_health = pd.DataFrame(health_details.data)
                if not df_health.empty:
                    df_health.to_excel(writer, sheet_name="Health Insurance Details", index=False)
                
                # Factory Insurance Details
                factory_details = self.supabase.table("factory_insurance_details").select("*").execute()
                df_factory = pd.DataFrame(factory_details.data)
                if not df_factory.empty:
                    df_factory.to_excel(writer, sheet_name="Factory Insurance Details", index=False)
                
                # Policy History
                policy_history = self.supabase.table("policy_history").select("*").execute()
                df_history = pd.DataFrame(policy_history.data)
                if not df_history.empty:
                    # Reorder columns for better readability
                    history_cols = df_history.columns.tolist()
                    
                    # Define preferred column order
                    preferred_order = [
                        'history_id', 'original_policy_id', 'client_id', 'member_id',
                        'insurance_company', 'product_name', 'policy_number',
                        'policy_from', 'policy_to', 'net_premium', 'gross_premium',
                        'sum_insured', 'agent_name', 'archived_at', 'archived_reason', 'archived_by'
                    ]
                    
                    # Reorder columns based on preference, keeping any extra columns at the end
                    ordered_cols = []
                    for col in preferred_order:
                        if col in history_cols:
                            ordered_cols.append(col)
                            history_cols.remove(col)
                    
                    # Add remaining columns
                    ordered_cols.extend(history_cols)
                    df_history = df_history[ordered_cols]
                    
                    # Format dates for better readability
                    date_columns = ['policy_from', 'policy_to', 'payment_date', 'archived_at', 'created_at', 'renewed_at']
                    for col in date_columns:
                        if col in df_history.columns:
                            df_history[col] = pd.to_datetime(df_history[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Format currency columns
                    currency_columns = ['net_premium', 'gross_premium', 'addon_premium', 'tp_tr_premium', 
                                      'commission_amount', 'sum_insured']
                    for col in currency_columns:
                        if col in df_history.columns:
                            df_history[col] = pd.to_numeric(df_history[col], errors='coerce').round(2)
                    
                    df_history.to_excel(writer, sheet_name="Policy History", index=False)
            
            # Apply formatting
            self._format_excel(self.local_excel_path)
            logger.info("Local Excel export completed successfully")
            
        except Exception as e:
            logger.error(f"Error exporting to local Excel: {e}")
            raise

    def export_to_drive(self):
        """Export current Supabase data to Google Drive Excel file"""
        with self.sync_lock:
            try:
                logger.info("Starting export to Google Drive...")
                
                # Create local Excel file with latest data
                self._export_supabase_to_local_excel()
                
                # Upload to Google Drive
                media = MediaFileUpload(
                    self.local_excel_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                self.drive_service.files().update(
                    fileId=self.excel_file_id,
                    media_body=media
                ).execute()
                
                logger.info("Excel file updated in Google Drive successfully")
                return True, "Excel file updated successfully in Google Drive"
                
            except Exception as e:
                logger.error(f"Error exporting to Drive: {e}")
                return False, f"Export failed: {str(e)}"

    def download_from_drive(self):
        """Download Excel file from Google Drive"""
        try:
            logger.info("Downloading Excel file from Google Drive...")
            
            request = self.drive_service.files().get_media(fileId=self.excel_file_id)
            file_io = io.BytesIO()
            downloader = MediaIoBaseDownload(file_io, request)
            
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            
            # Save to local file
            with open(self.local_excel_path, 'wb') as f:
                f.write(file_io.getvalue())
            
            logger.info("Excel file downloaded successfully")
            return True, "Excel file downloaded successfully"
            
        except Exception as e:
            logger.error(f"Error downloading from Drive: {e}")
            return False, f"Download failed: {str(e)}"

    def get_drive_file_info(self):
        """Get information about the Excel file in Google Drive"""
        try:
            file_info = self.drive_service.files().get(
                fileId=self.excel_file_id,
                fields='id,name,modifiedTime,size,webViewLink,webContentLink'
            ).execute()
            
            return {
                'file_id': file_info.get('id'),
                'name': file_info.get('name'),
                'modified_time': file_info.get('modifiedTime'),
                'size': file_info.get('size'),
                'view_link': file_info.get('webViewLink'),
                'download_link': file_info.get('webContentLink')
            }
            
        except Exception as e:
            logger.error(f"Error getting file info: {e}")
            return None

    def get_shareable_link(self):
        """Get shareable Google Drive link for the Excel file"""
        try:
            file_info = self.get_drive_file_info()
            if file_info:
                return file_info['view_link']
            return None
        except Exception as e:
            logger.error(f"Error getting shareable link: {e}")
            return None

    def export_policy_history_report(self, policy_id=None, client_id=None, date_from=None, date_to=None):
        """
        Create a detailed policy history report with enhanced formatting
        
        Args:
            policy_id (int): Specific policy ID to filter by (optional)
            client_id (str): Specific client ID to filter by (optional)
            date_from (str): Start date for archived_at filter (optional)
            date_to (str): End date for archived_at filter (optional)
        
        Returns:
            str: Path to the generated Excel file
        """
        try:
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"policy_history_report_{timestamp}.xlsx"
            report_path = os.path.join(os.getcwd(), filename)
            
            logger.info(f"Generating policy history report: {filename}")
            
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # Build query for policy history
                query = self.supabase.table("policy_history").select("*")
                
                # Apply filters
                if policy_id:
                    query = query.eq("original_policy_id", policy_id)
                if client_id:
                    query = query.eq("client_id", client_id)
                if date_from:
                    query = query.gte("archived_at", date_from)
                if date_to:
                    query = query.lte("archived_at", date_to)
                
                # Execute query
                history_result = query.order("archived_at", desc=True).execute()
                df_history = pd.DataFrame(history_result.data)
                
                if not df_history.empty:
                    # Enhanced column ordering and formatting
                    preferred_order = [
                        'history_id', 'original_policy_id', 'client_id', 'member_id',
                        'insurance_company', 'product_name', 'policy_number',
                        'policy_from', 'policy_to', 'net_premium', 'gross_premium', 'sum_insured',
                        'agent_name', 'business_type', 'group_name', 'subgroup_name',
                        'commission_percentage', 'commission_amount', 'payment_date',
                        'archived_at', 'archived_reason', 'archived_by', 'remarks'
                    ]
                    
                    # Reorder columns
                    available_cols = df_history.columns.tolist()
                    ordered_cols = [col for col in preferred_order if col in available_cols]
                    remaining_cols = [col for col in available_cols if col not in ordered_cols]
                    final_cols = ordered_cols + remaining_cols
                    
                    df_history = df_history[final_cols]
                    
                    # Format data for better readability
                    # Format dates
                    date_columns = ['policy_from', 'policy_to', 'payment_date', 'archived_at', 'created_at', 'renewed_at']
                    for col in date_columns:
                        if col in df_history.columns:
                            df_history[col] = pd.to_datetime(df_history[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Format currency columns
                    currency_columns = ['net_premium', 'gross_premium', 'addon_premium', 'tp_tr_premium', 
                                      'commission_amount', 'sum_insured']
                    for col in currency_columns:
                        if col in df_history.columns:
                            df_history[col] = pd.to_numeric(df_history[col], errors='coerce').round(2)
                    
                    # Add summary information
                    summary_data = {
                        'Total Historical Records': len(df_history),
                        'Unique Policies': df_history['original_policy_id'].nunique() if 'original_policy_id' in df_history.columns else 0,
                        'Unique Clients': df_history['client_id'].nunique() if 'client_id' in df_history.columns else 0,
                        'Date Range': f"{df_history['archived_at'].min()} to {df_history['archived_at'].max()}" if 'archived_at' in df_history.columns else 'N/A',
                        'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Create summary sheet
                    df_summary = pd.DataFrame(list(summary_data.items()), columns=['Metric', 'Value'])
                    df_summary.to_excel(writer, sheet_name="Summary", index=False)
                    
                    # Main history data
                    df_history.to_excel(writer, sheet_name="Policy History Details", index=False)
                    
                    # Group by policy for analysis
                    if 'original_policy_id' in df_history.columns:
                        policy_summary = df_history.groupby('original_policy_id').agg({
                            'history_id': 'count',
                            'archived_at': ['min', 'max'],
                            'net_premium': ['first', 'last'] if 'net_premium' in df_history.columns else 'count',
                            'insurance_company': lambda x: ' → '.join(x.unique()) if len(x.unique()) > 1 else x.iloc[0]
                        }).round(2)
                        
                        # Flatten column names
                        policy_summary.columns = ['_'.join(col).strip() if isinstance(col, tuple) else col for col in policy_summary.columns]
                        policy_summary = policy_summary.reset_index()
                        
                        # Rename columns for clarity
                        column_mapping = {
                            'history_id_count': 'Total_Versions',
                            'archived_at_min': 'First_Archived',
                            'archived_at_max': 'Last_Archived',
                            'net_premium_first': 'Original_Premium',
                            'net_premium_last': 'Latest_Premium',
                            'insurance_company_<lambda>': 'Company_Changes'
                        }
                        
                        for old_name, new_name in column_mapping.items():
                            if old_name in policy_summary.columns:
                                policy_summary.rename(columns={old_name: new_name}, inplace=True)
                        
                        policy_summary.to_excel(writer, sheet_name="Policy Summary", index=False)
                
                else:
                    # Create empty sheet with message
                    df_empty = pd.DataFrame({'Message': ['No policy history records found with the specified criteria']})
                    df_empty.to_excel(writer, sheet_name="Policy History Details", index=False)
            
            # Apply enhanced formatting
            self._format_excel(report_path)
            logger.info(f"Policy history report generated successfully: {filename}")
            
            return report_path
            
        except Exception as e:
            logger.error(f"Error generating policy history report: {e}")
            raise


# Global instance for use across the application
excel_service = ExcelSyncService()
