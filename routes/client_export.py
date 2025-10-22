from flask import Blueprint, request, jsonify, send_file, flash, redirect, url_for
from flask_login import login_required
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client
from dynamic_config import Config

client_export_bp = Blueprint('client_export', __name__)

# Initialize Supabase
supabase = create_client(Config.SUPABASE_URL, Config.SUPABASE_KEY)

def convert_date_for_display(date_str):
    """Convert database date to display format"""
    if not date_str:
        return ""
    try:
        if isinstance(date_str, str):
            # Handle different date formats
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    return date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    continue
        return str(date_str)
    except:
        return str(date_str) if date_str else ""

@client_export_bp.route('/export_client_data/<client_id>')
@login_required
def export_client_data(client_id):
    """Export all client policy data to Excel file"""
    try:
        # Get client information
        client_result = supabase.table("clients").select("*").eq("client_id", client_id).single().execute()
        if not client_result.data:
            flash("Client not found", "error")
            return redirect(url_for('existing_policies.list_all'))
        
        client = client_result.data
        
        # Get all active policies for this client
        policies_result = supabase.table("policies").select("""
            policy_id, policy_number, insurance_company, product_name, agent_name,
            policy_from, policy_to, payment_date, business_type, group_name, subgroup_name,
            remarks, sum_insured, net_premium, addon_premium, tp_tr_premium, gst_percentage, 
            gross_premium
        """).eq("client_id", client_id).execute()
        
        policies = policies_result.data
        
        if not policies:
            flash("No active policies found for this client", "warning")
            return redirect(url_for('existing_policies.list_all'))
        
        # Get health insurance details for all policies
        health_details = {}
        health_members = {}
        
        for policy in policies:
            policy_id = policy['policy_id']
            
            # Get health insurance details
            health_result = supabase.table("health_insurance_details").select("*").eq("policy_id", policy_id).execute()
            if health_result.data:
                health_details[policy_id] = health_result.data[0]
                
                # Get health members
                health_id = health_result.data[0]['health_id']
                members_result = supabase.table("health_insured_members").select("*").eq("health_id", health_id).execute()
                health_members[policy_id] = members_result.data
        
        # Get factory insurance details for all policies
        factory_details = {}
        for policy in policies:
            policy_id = policy['policy_id']
            factory_result = supabase.table("factory_insurance_details").select("*").eq("policy_id", policy_id).execute()
            if factory_result.data:
                factory_details[policy_id] = factory_result.data[0]
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = f"{client_id} Policy Data"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Determine maximum number of health members across all policies
        max_health_members = 0
        for policy_id in health_members:
            max_health_members = max(max_health_members, len(health_members[policy_id]))
        
        # Create headers (removed commission fields for client privacy)
        headers = [
            "Policy Number", "Insurance Company", "Product Type", "Agent Name",
            "Policy Start Date", "Policy End Date", "Payment Date", "Business Type",
            "Group", "Subgroup", "Remarks", "Sum Insured", "Net Premium/OD", 
            "Addon Premium", "TP/TR Premium", "GST %", "Gross Premium"
        ]
        
        # Add health insurance headers if applicable
        if max_health_members > 0:
            headers.extend(["Health Plan Type", "Floater Sum Insured", "Floater Bonus", "Floater Deductible"])
            for i in range(max_health_members):
                member_num = i + 1
                headers.extend([
                    f"Member {member_num} Name",
                    f"Member {member_num} Sum Insured", 
                    f"Member {member_num} Bonus",
                    f"Member {member_num} Deductible"
                ])
        
        # Add factory insurance headers
        factory_headers = [
            "Building Coverage", "Plant & Machinery Coverage", 
            "Furniture & Fittings Coverage", "Stocks Coverage", 
            "Electrical Installations Coverage"
        ]
        headers.extend(factory_headers)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data rows
        for row_idx, policy in enumerate(policies, 2):
            policy_id = policy['policy_id']
            
            # Basic policy data (commission information excluded for client privacy)
            
            row_data = [
                policy.get('policy_number', ''),
                policy.get('insurance_company', ''),
                policy.get('product_name', ''),
                policy.get('agent_name', ''),
                convert_date_for_display(policy.get('policy_from')),
                convert_date_for_display(policy.get('policy_to')),
                convert_date_for_display(policy.get('payment_date')),
                policy.get('business_type', ''),
                policy.get('group_name', ''),
                policy.get('subgroup_name', ''),
                policy.get('remarks', ''),
                policy.get('sum_insured', ''),
                policy.get('net_premium', ''),
                policy.get('addon_premium', ''),
                policy.get('tp_tr_premium', ''),
                policy.get('gst_percentage', ''),
                policy.get('gross_premium', '')
            ]
            
            # Add health insurance data if applicable
            if max_health_members > 0:
                if policy_id in health_details:
                    health_detail = health_details[policy_id]
                    row_data.extend([
                        health_detail.get('plan_type', ''),
                        health_detail.get('floater_sum_insured', ''),
                        health_detail.get('floater_bonus', ''),
                        health_detail.get('floater_deductible', '')
                    ])
                    
                    # Add member data
                    members = health_members.get(policy_id, [])
                    for i in range(max_health_members):
                        if i < len(members):
                            member = members[i]
                            row_data.extend([
                                member.get('member_name', ''),
                                member.get('sum_insured', ''),
                                member.get('bonus', ''),
                                member.get('deductible', '')
                            ])
                        else:
                            row_data.extend(['', '', '', ''])  # Empty cells for missing members
                else:
                    # No health insurance for this policy
                    row_data.extend([''] * (4 + max_health_members * 4))
            
            # Add factory insurance data
            if policy_id in factory_details:
                factory_detail = factory_details[policy_id]
                row_data.extend([
                    factory_detail.get('building', ''),
                    factory_detail.get('plant_machinery', ''),
                    factory_detail.get('furniture_fittings', ''),
                    factory_detail.get('stocks', ''),
                    factory_detail.get('electrical_installations', '')
                ])
            else:
                row_data.extend(['', '', '', '', ''])  # Empty factory columns
            
            # Write row data
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col > 11 and value:  # Format numeric columns
                    try:
                        if float(value):
                            cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        pass
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Generate filename
        filename = f"{client_id}_data.xlsx"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error exporting client data: {e}")
        flash(f"Error exporting data: {str(e)}", "error")
        return redirect(url_for('existing_policies.list_all'))
    
    finally:
        # Clean up temporary file
        try:
            if 'temp_file' in locals():
                os.unlink(temp_file.name)
        except:
            pass
