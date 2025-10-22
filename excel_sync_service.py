"""
Real-time Excel Sync Service for Insurance Portal
Based on the original RealtimeSupabaseExcelSync implementation
"""

import os
import time
import threading
import logging
from datetime import datetime, date
from pathlib import Path
from supabase import create_client, Client
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import hashlib
import json
from dynamic_config import Config
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io

logger = logging.getLogger(__name__)

class RealtimeExcelSync:
    def __init__(self, excel_file="insurance_data.xlsx"):
        """Initialize real-time sync between Supabase and Excel"""
        self.supabase: Client = create_client(Config.SUPABASE_URL, Config.SUPABASE_KEY)
        self.excel_file = excel_file
        self.local_excel_path = os.path.join(os.getcwd(), excel_file)
        self.is_syncing = False
        self.last_excel_hash = {}
        self.last_supabase_data = {}
        self.sync_lock = threading.Lock()
        self.sync_thread = None
        self.stop_sync = False
        
        # Initialize Google Drive
        self.drive_service = self._init_google_drive()
        self.drive_file_id = None
        
        # Create initial Excel file if it doesn't exist
        if not os.path.exists(self.local_excel_path):
            self._initial_export()
        
        # Setup Google Drive file
        self._setup_drive_file()
        
        self._update_hashes()
        logger.info(f"Real-time Excel sync initialized for: {excel_file}")
    
    def _determine_financial_year(self, policy_from_date):
        """Determine financial year from policy start date (April to March)"""
        if not policy_from_date:
            return None
        
        try:
            # Handle different date formats
            if isinstance(policy_from_date, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
                    try:
                        start_date = datetime.strptime(policy_from_date, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    return None
            elif isinstance(policy_from_date, datetime):
                start_date = policy_from_date.date()
            elif isinstance(policy_from_date, date):
                start_date = policy_from_date
            else:
                return None
            
            # Financial year logic: April to March
            if start_date.month >= 4:  # April to December
                return f"{start_date.year}-{str(start_date.year + 1)[2:]}"
            else:  # January to March
                return f"{start_date.year - 1}-{str(start_date.year)[2:]}"
                
        except Exception as e:
            logger.warning(f"Error determining financial year for date {policy_from_date}: {e}")
            return None
    
    def _get_policies_with_insurance_details(self):
        """Get all policies with health and factory insurance details"""
        try:
            # Get all policies with client and member info
            policies = self.supabase.table("policies").select(
                "*, clients(name, phone, email), members(member_name)"
            ).execute()
            
            policies_data = policies.data
            
            # Get health insurance details
            health_details = {}
            health_members = {}
            
            for policy in policies_data:
                policy_id = policy['policy_id']
                
                # Get health insurance details
                health_result = self.supabase.table("health_insurance_details").select("*").eq("policy_id", policy_id).execute()
                if health_result.data:
                    health_details[policy_id] = health_result.data[0]
                    
                    # Get health members
                    health_id = health_result.data[0]['health_id']
                    members_result = self.supabase.table("health_insured_members").select("*").eq("health_id", health_id).execute()
                    health_members[policy_id] = members_result.data
            
            # Get factory insurance details
            factory_details = {}
            for policy in policies_data:
                policy_id = policy['policy_id']
                factory_result = self.supabase.table("factory_insurance_details").select("*").eq("policy_id", policy_id).execute()
                if factory_result.data:
                    factory_details[policy_id] = factory_result.data[0]
            
            return policies_data, health_details, health_members, factory_details
            
        except Exception as e:
            logger.error(f"Error getting policies with insurance details: {e}")
            return [], {}, {}, {}
    
    def _get_claims_with_details(self):
        """Get all claims with policy and client information"""
        try:
            # Get all claims with policy and client info
            claims = self.supabase.table("claims").select(
                "*, policies(policy_number, policy_from, clients(name, phone, email))"
            ).execute()
            
            return claims.data
            
        except Exception as e:
            logger.error(f"Error getting claims with details: {e}")
            return []
    
    def _convert_date_for_display(self, date_str):
        """Convert database date to display format"""
        if not date_str:
            return ""
        try:
            if isinstance(date_str, str):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
                    try:
                        date_obj = datetime.strptime(date_str, fmt)
                        return date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        continue
            return str(date_str)
        except:
            return str(date_str) if date_str else ""

    def _init_google_drive(self):
        """Initialize Google Drive API service (same as existing integration)"""
        try:
            credentials = Credentials.from_service_account_file(
                Config.GOOGLE_CREDENTIALS_FILE,
                scopes=['https://www.googleapis.com/auth/drive']
            )
            service = build('drive', 'v3', credentials=credentials)
            logger.info("Google Drive service initialized for shared drives")
            return service
        except Exception as e:
            logger.error(f"Failed to initialize Google Drive: {e}")
            return None

    def _setup_drive_file(self):
        """Find or create Excel file in Google Drive (using shared drive approach)"""
        if not self.drive_service:
            logger.warning("Google Drive service not available")
            return
            
        try:
            # Search for existing file in root folder (ID: 0AOc3bRLhlrgzUk9PVA)
            ROOT_FOLDER_ID = "0AOc3bRLhlrgzUk9PVA"
            query = f"name='{self.excel_file}' and '{ROOT_FOLDER_ID}' in parents and trashed=false"
            results = self.drive_service.files().list(
                q=query,
                fields="files(id, name, parents, webViewLink)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True
            ).execute()
            files = results.get('files', [])
            
            if files:
                self.drive_file_id = files[0]['id']
                logger.info(f"Found existing Excel file in shared drive: {self.drive_file_id}")
            else:
                # Create new file using archive folder as parent (same as existing pattern)
                self._upload_to_drive()
                logger.info(f"Created new Excel file in shared drive: {self.drive_file_id}")
                
        except Exception as e:
            logger.error(f"Error setting up Drive file: {e}")
            self.drive_file_id = None

    def _upload_to_drive(self):
        """Upload Excel file to Google Drive (using shared drive approach like existing code)"""
        if not self.drive_service:
            return
            
        try:
            # Use root folder as parent (ID: 0AOc3bRLhlrgzUk9PVA)
            ROOT_FOLDER_ID = "0AOc3bRLhlrgzUk9PVA"
            file_metadata = {
                'name': self.excel_file,
                'parents': [ROOT_FOLDER_ID]  # Use root folder for Excel file
            }
            
            media = MediaFileUpload(
                self.local_excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            file = self.drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, name, webViewLink',
                supportsAllDrives=True  # Essential for shared drives
            ).execute()
            
            self.drive_file_id = file.get('id')
            
            # Set permissions for shared drive access (same as existing pattern)
            try:
                permission = {
                    'type': 'anyone',
                    'role': 'reader'  # Reader access for shared drives
                }
                self.drive_service.permissions().create(
                    fileId=self.drive_file_id,
                    body=permission,
                    supportsAllDrives=True
                ).execute()
                logger.info("Excel file permissions set for shared drive access")
            except Exception as perm_error:
                logger.warning(f"Could not set permissions (may already be inherited): {perm_error}")
            
            logger.info(f"Excel file uploaded to root folder in shared drive: {self.drive_file_id}")
            
        except Exception as e:
            logger.error(f"Error uploading to shared drive: {e}")
            self.drive_file_id = None

    def _update_drive_file(self):
        """Update existing file in Google Drive (using shared drive approach)"""
        if not self.drive_service or not self.drive_file_id:
            return
            
        try:
            media = MediaFileUpload(
                self.local_excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            self.drive_service.files().update(
                fileId=self.drive_file_id,
                media_body=media,
                supportsAllDrives=True  # Essential for shared drives
            ).execute()
            
            logger.info("Excel file updated in shared drive")
            
        except Exception as e:
            logger.error(f"Error updating shared drive file: {e}")

    def _initial_export(self):
        """Create initial Excel file from Supabase"""
        logger.info("Creating initial Excel file...")
        self.export_supabase_to_excel()

    def _get_file_hash(self, filepath):
        """Calculate hash of Excel file"""
        if not os.path.exists(filepath):
            return None
        return hashlib.md5(open(filepath, 'rb').read()).hexdigest()

    def _get_data_hash(self, data):
        """Calculate hash of data"""
        return hashlib.md5(json.dumps(data, sort_keys=True, default=str).encode()).hexdigest()

    def _update_hashes(self):
        """Update stored hashes"""
        self.last_excel_hash = self._get_file_hash(self.local_excel_path)

        # Get current Supabase data hashes
        try:
            clients = self.supabase.table("clients").select("*").execute()
            members = self.supabase.table("members").select("*").execute()
            policies = self.supabase.table("policies").select("*").execute()
            pending = self.supabase.table("pending_policies").select("*").execute()
            claims = self.supabase.table("claims").select("*").execute()
            policy_history = self.supabase.table("policy_history").select("*").execute()

            self.last_supabase_data = {
                'clients': self._get_data_hash(clients.data),
                'members': self._get_data_hash(members.data),
                'policies': self._get_data_hash(policies.data),
                'pending_policies': self._get_data_hash(pending.data),
                'claims': self._get_data_hash(claims.data),
                'policy_history': self._get_data_hash(policy_history.data)
            }
        except Exception as e:
            logger.error(f"Error updating Supabase hashes: {e}")

    def _format_excel(self, filename):
        """Apply formatting to Excel file"""
        try:
            wb = load_workbook(filename)

            for sheet in wb.worksheets:
                # Format headers
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
        except Exception as e:
            logger.warning(f"Could not format Excel: {e}")

    def export_supabase_to_excel(self):
        """Export Supabase data to Excel with financial year-based sheets"""
        with self.sync_lock:
            logger.info(f"[{datetime.now().strftime('%H:%M:%S')}] Exporting Supabase → Excel (Financial Year Sheets)...")

            try:
                # Create new workbook
                wb = Workbook()
                
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
                
                # Get policies with insurance details
                policies_data, health_details, health_members, factory_details = self._get_policies_with_insurance_details()
                
                if not policies_data:
                    logger.warning("No policies found to export")
                    # Create a placeholder sheet
                    ws = wb.create_sheet("No Data")
                    ws['A1'] = "No policies found"
                    wb.save(self.local_excel_path)
                    return
                
                # Group policies by financial year based on policy start date
                policies_by_year = {}
                for policy in policies_data:
                    financial_year = self._determine_financial_year(policy.get('policy_from'))
                    if financial_year:
                        if financial_year not in policies_by_year:
                            policies_by_year[financial_year] = []
                        policies_by_year[financial_year].append(policy)
                
                # Get claims data
                claims_data = self._get_claims_with_details()
                
                # Group claims by financial year based on claim creation date
                claims_by_year = {}
                for claim in claims_data:
                    # Use claim creation date instead of policy start date
                    claim_date = claim.get('created_at')
                    if claim_date:
                        financial_year = self._determine_financial_year(claim_date)
                        if financial_year:
                            if financial_year not in claims_by_year:
                                claims_by_year[financial_year] = []
                            claims_by_year[financial_year].append(claim)
                
                # Create sheets for each financial year (policies)
                for financial_year in sorted(policies_by_year.keys(), reverse=True):
                    year_policies = policies_by_year[financial_year]
                    self._create_financial_year_sheet(wb, financial_year, year_policies, health_details, health_members, factory_details)
                
                # Create sheets for each financial year (claims)
                for financial_year in sorted(claims_by_year.keys(), reverse=True):
                    year_claims = claims_by_year[financial_year]
                    self._create_claims_financial_year_sheet(wb, financial_year, year_claims)
                
                # Also create summary sheets
                self._create_clients_sheet(wb)
                self._create_members_sheet(wb)
                self._create_pending_policies_sheet(wb)
                self._create_policy_history_sheet(wb)
                
                # Save workbook
                wb.save(self.local_excel_path)
                
                # Apply additional formatting
                self._format_excel(self.local_excel_path)
                
                # Upload to Google Drive
                self._update_drive_file()
                self._update_hashes()
                
                logger.info(f"[{datetime.now().strftime('%H:%M:%S')}] ✓ Export completed with {len(policies_by_year)} financial year sheets")
                
            except Exception as e:
                logger.error(f"[{datetime.now().strftime('%H:%M:%S')}] ✗ Export error: {e}")
    
    def _create_financial_year_sheet(self, workbook, financial_year, policies, health_details, health_members, factory_details):
        """Create a sheet for a specific financial year with all policy data including health and factory details"""
        try:
            ws = workbook.create_sheet(financial_year)
            
            # Determine maximum number of health members across all policies
            max_health_members = 0
            for policy in policies:
                policy_id = policy['policy_id']
                if policy_id in health_members:
                    max_health_members = max(max_health_members, len(health_members[policy_id]))
            
            # Define headers with client-friendly names
            headers = [
                "Policy ID", "Client Name", "Member Name", "Policy Number", "Insurance Company", 
                "Product Type", "Agent Name", "Policy Start Date", "Policy End Date", "Payment Date",
                "Business Type", "Group", "Subgroup", "Remarks", "Sum Insured", "Net Premium", 
                "Addon Premium", "TP/TR Premium", "GST %", "Gross Premium", "Commission %", "Commission Amount", 
                "Commission Received", "One Time Insurance", "Payment Details", "File Path", "Drive File ID", 
                "Drive Path", "Drive URL", "Last Reminder Sent", "Renewed At", "Created At", "Updated At"
            ]
            
            # Add health insurance headers if applicable
            if max_health_members > 0:
                headers.extend(["Health Plan Type", "Floater Sum Insured", "Floater Bonus", "Floater Deductible"])
                for i in range(max_health_members):
                    member_num = i + 1
                    headers.extend([
                        f"Health Member {member_num} Name",
                        f"Health Member {member_num} Sum Insured", 
                        f"Health Member {member_num} Bonus",
                        f"Health Member {member_num} Deductible"
                    ])
            
            # Add factory insurance headers
            factory_headers = [
                "Factory Building Coverage", "Factory Plant & Machinery Coverage", 
                "Factory Furniture & Fittings Coverage", "Factory Stocks Coverage", 
                "Factory Electrical Installations Coverage"
            ]
            headers.extend(factory_headers)
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, policy in enumerate(policies, 2):
                policy_id = policy['policy_id']
                
                # Basic policy data
                client_info = policy.get('clients', {})
                member_info = policy.get('members', {})
                
                # Calculate commission amount
                commission_amount = ''
                try:
                    net_premium = policy.get('net_premium')
                    commission_percentage = policy.get('commission_percentage')
                    if net_premium and commission_percentage:
                        commission_amount = float(net_premium) * float(commission_percentage) / 100
                        commission_amount = f"{commission_amount:.2f}"
                except (ValueError, TypeError):
                    commission_amount = ''
                
                row_data = [
                    policy.get('policy_id', ''),
                    client_info.get('name', '') if client_info else '',
                    member_info.get('member_name', '') if member_info else '',
                    policy.get('policy_number', ''),
                    policy.get('insurance_company', ''),
                    policy.get('product_name', ''),
                    policy.get('agent_name', ''),
                    self._convert_date_for_display(policy.get('policy_from')),
                    self._convert_date_for_display(policy.get('policy_to')),
                    self._convert_date_for_display(policy.get('payment_date')),
                    policy.get('business_type', ''),
                    policy.get('group_name', ''),
                    policy.get('subgroup_name', ''),
                    policy.get('remarks', ''),
                    policy.get('sum_insured', ''),
                    policy.get('net_premium', ''),
                    policy.get('addon_premium', ''),
                    policy.get('tp_tr_premium', ''),
                    policy.get('gst_percentage', ''),
                    policy.get('gross_premium', ''),
                    policy.get('commission_percentage', ''),
                    commission_amount,
                    'Yes' if policy.get('commission_received') else 'No',
                    'Yes' if policy.get('one_time_insurance') else 'No',
                    policy.get('payment_details', ''),
                    policy.get('file_path', ''),
                    policy.get('drive_file_id', ''),
                    policy.get('drive_path', ''),
                    policy.get('drive_url', ''),
                    self._convert_date_for_display(policy.get('last_reminder_sent')),
                    self._convert_date_for_display(policy.get('renewed_at')),
                    self._convert_date_for_display(policy.get('created_at')),
                    self._convert_date_for_display(policy.get('updated_at'))
                ]
                
                # Add health insurance data if applicable
                if max_health_members > 0:
                    if policy_id in health_details:
                        health_detail = health_details[policy_id]
                        row_data.extend([
                            health_detail.get('plan_type', ''),
                            health_detail.get('floater_sum_insured', ''),
                            health_detail.get('floater_bonus', ''),
                            health_detail.get('floater_deductible', '')
                        ])
                        
                        # Add member data
                        members = health_members.get(policy_id, [])
                        for i in range(max_health_members):
                            if i < len(members):
                                member = members[i]
                                row_data.extend([
                                    member.get('member_name', ''),
                                    member.get('sum_insured', ''),
                                    member.get('bonus', ''),
                                    member.get('deductible', '')
                                ])
                            else:
                                row_data.extend(['', '', '', ''])  # Empty cells for missing members
                    else:
                        # No health insurance for this policy
                        row_data.extend([''] * (4 + max_health_members * 4))
                
                # Add factory insurance data
                if policy_id in factory_details:
                    factory_detail = factory_details[policy_id]
                    row_data.extend([
                        factory_detail.get('building', ''),
                        factory_detail.get('plant_machinery', ''),
                        factory_detail.get('furniture_fittings', ''),
                        factory_detail.get('stocks', ''),
                        factory_detail.get('electrical_installations', '')
                    ])
                else:
                    row_data.extend(['', '', '', '', ''])  # Empty factory columns
                
                # Write row data
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Created financial year sheet '{financial_year}' with {len(policies)} policies")
            
        except Exception as e:
            logger.error(f"Error creating financial year sheet {financial_year}: {e}")
    
    def _create_claims_financial_year_sheet(self, workbook, financial_year, claims):
        """Create a claims sheet for a specific financial year"""
        try:
            ws = workbook.create_sheet(f"Claims {financial_year}")
            
            # Define claims headers
            headers = [
                "Claim ID", "Policy Number", "Client Name", "Member Name", "Claim Type", 
                "Claim Number", "Diagnosis", "Hospital Name", "Admission Date", "Discharge Date",
                "Claimed Amount", "Approved Amount", "Settled Amount", "Status", "Settlement Date",
                "UTR No", "Remarks", "Created At", "Updated At"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, claim in enumerate(claims, 2):
                policy_info = claim.get('policies', {})
                client_info = policy_info.get('clients', {}) if policy_info else {}
                
                row_data = [
                    claim.get('claim_id', ''),
                    policy_info.get('policy_number', '') if policy_info else '',
                    client_info.get('name', '') if client_info else '',
                    claim.get('member_name', ''),
                    claim.get('claim_type', ''),
                    claim.get('claim_number', ''),
                    claim.get('diagnosis', ''),
                    claim.get('hospital_name', ''),
                    self._convert_date_for_display(claim.get('admission_date')),
                    self._convert_date_for_display(claim.get('discharge_date')),
                    claim.get('claimed_amount', ''),
                    claim.get('approved_amount', ''),
                    claim.get('settled_amount', ''),
                    claim.get('status', ''),
                    self._convert_date_for_display(claim.get('settlement_date')),
                    claim.get('utr_no', ''),
                    claim.get('remarks', ''),
                    self._convert_date_for_display(claim.get('created_at')),
                    self._convert_date_for_display(claim.get('updated_at'))
                ]
                
                # Write row data
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Created claims financial year sheet 'Claims {financial_year}' with {len(claims)} claims")
            
        except Exception as e:
            logger.error(f"Error creating claims financial year sheet {financial_year}: {e}")
    
    def _create_clients_sheet(self, workbook):
        """Create clients summary sheet"""
        try:
            clients = self.supabase.table("clients").select("*").execute()
            df_clients = pd.DataFrame(clients.data)
            if not df_clients.empty:
                # Convert DataFrame to sheet
                ws = workbook.create_sheet("Clients")
                
                # Write headers
                for col, header in enumerate(df_clients.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Write data
                for row_idx, (_, row) in enumerate(df_clients.iterrows(), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        except Exception as e:
            logger.error(f"Error creating clients sheet: {e}")
    
    def _create_members_sheet(self, workbook):
        """Create members summary sheet"""
        try:
            members = self.supabase.table("members").select(
                "*, clients(name, phone, email)"
            ).execute()
            df_members = pd.DataFrame(members.data)
            if not df_members.empty:
                # Process client join data
                if 'clients' in df_members.columns:
                    client_df = pd.json_normalize(df_members['clients'])
                    client_df.columns = ['client_' + col for col in client_df.columns]
                    df_members = pd.concat([df_members.drop('clients', axis=1), client_df], axis=1)
                
                # Convert DataFrame to sheet
                ws = workbook.create_sheet("Members")
                
                # Write headers
                for col, header in enumerate(df_members.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Write data
                for row_idx, (_, row) in enumerate(df_members.iterrows(), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        except Exception as e:
            logger.error(f"Error creating members sheet: {e}")
    
    def _create_pending_policies_sheet(self, workbook):
        """Create pending policies summary sheet"""
        try:
            pending = self.supabase.table("pending_policies").select(
                "*, clients(name, phone, email), members(member_name)"
            ).execute()
            df_pending = pd.DataFrame(pending.data)
            if not df_pending.empty:
                # Process join data
                if 'clients' in df_pending.columns:
                    client_df = pd.json_normalize(df_pending['clients'])
                    client_df.columns = ['client_' + col for col in client_df.columns]
                    df_pending = pd.concat([df_pending.drop('clients', axis=1), client_df], axis=1)
                
                if 'members' in df_pending.columns:
                    member_df = pd.json_normalize(df_pending['members'])
                    member_df.columns = ['member_' + col for col in member_df.columns]
                    df_pending = pd.concat([df_pending.drop('members', axis=1), member_df], axis=1)
                
                # Convert DataFrame to sheet
                ws = workbook.create_sheet("Pending Policies")
                
                # Write headers
                for col, header in enumerate(df_pending.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Write data
                for row_idx, (_, row) in enumerate(df_pending.iterrows(), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        except Exception as e:
            logger.error(f"Error creating pending policies sheet: {e}")

    def _create_policy_history_sheet(self, workbook):
        """Create policy history sheet with additional details (factory, health insurance)"""
        try:
            policy_history = self.supabase.table("policy_history").select("*").execute()
            if not policy_history.data:
                logger.info("No policy history data found")
                return
            
            # Get factory and health insurance history details
            factory_history = {}
            health_history = {}
            health_members_history = {}
            
            try:
                factory_data = self.supabase.table("policy_history_factory_details").select("*").execute()
                for item in factory_data.data:
                    factory_history[item['history_id']] = item
            except Exception as e:
                logger.warning(f"Could not fetch factory history details: {e}")
            
            try:
                health_data = self.supabase.table("policy_history_health_details").select("*").execute()
                for item in health_data.data:
                    health_history[item['history_id']] = item
                    
                # Get health members
                health_members_data = self.supabase.table("policy_history_health_members").select("*").execute()
                for member in health_members_data.data:
                    health_id = member['history_health_id']
                    if health_id not in health_members_history:
                        health_members_history[health_id] = []
                    health_members_history[health_id].append(member)
            except Exception as e:
                logger.warning(f"Could not fetch health history details: {e}")
            
            # Enrich policy history with additional details
            enriched_history = []
            for record in policy_history.data:
                history_id = record['history_id']
                enriched_record = record.copy()
                
                # Add factory details if available
                if history_id in factory_history:
                    factory = factory_history[history_id]
                    enriched_record['factory_building'] = factory.get('building')
                    enriched_record['factory_plant_machinery'] = factory.get('plant_machinery')
                    enriched_record['factory_furniture_fittings'] = factory.get('furniture_fittings')
                    enriched_record['factory_stocks'] = factory.get('stocks')
                    enriched_record['factory_electrical_installations'] = factory.get('electrical_installations')
                
                # Add health details if available
                if history_id in health_history:
                    health = health_history[history_id]
                    enriched_record['health_plan_type'] = health.get('plan_type')
                    enriched_record['health_floater_sum_insured'] = health.get('floater_sum_insured')
                    enriched_record['health_floater_bonus'] = health.get('floater_bonus')
                    enriched_record['health_floater_deductible'] = health.get('floater_deductible')
                    
                    # Add health members as comma-separated string
                    health_id = health.get('history_health_id')
                    if health_id and health_id in health_members_history:
                        members = health_members_history[health_id]
                        member_details = []
                        for m in members:
                            details = f"{m['member_name']} (SI: {m.get('sum_insured', 'N/A')}, Bonus: {m.get('bonus', 'N/A')})"
                            member_details.append(details)
                        enriched_record['health_members'] = '; '.join(member_details)
                
                enriched_history.append(enriched_record)
                
            df_history = pd.DataFrame(enriched_history)
            
            # Convert DataFrame to sheet
            ws = workbook.create_sheet("Policy History")
            
            # Define user-friendly column headers
            column_mapping = {
                'history_id': 'History ID',
                'original_policy_id': 'Original Policy ID',
                'client_id': 'Client ID',
                'member_id': 'Member ID',
                'insurance_company': 'Insurance Company',
                'product_name': 'Product Name',
                'policy_number': 'Policy Number',
                'one_time_insurance': 'One Time Insurance',
                'commission_received': 'Commission Received',
                'file_path': 'File Path',
                'drive_file_id': 'Drive File ID',
                'drive_path': 'Drive Path',
                'drive_url': 'Drive URL',
                'payment_date': 'Payment Date',
                'agent_name': 'Agent Name',
                'policy_from': 'Policy Start Date',
                'policy_to': 'Policy End Date',
                'payment_details': 'Payment Details',
                'net_premium': 'Net Premium',
                'addon_premium': 'Addon Premium',
                'tp_tr_premium': 'TP/TR Premium',
                'gst_percentage': 'GST %',
                'gross_premium': 'Gross Premium',
                'commission_percentage': 'Commission %',
                'commission_amount': 'Commission Amount',
                'business_type': 'Business Type',
                'group_name': 'Group Name',
                'subgroup_name': 'Subgroup Name',
                'remarks': 'Remarks',
                'sum_insured': 'Sum Insured',
                'last_reminder_sent': 'Last Reminder Sent',
                'renewed_at': 'Renewed At',
                'created_at': 'Created At',
                'updated_at': 'Updated At',
                'archived_at': 'Archived At',
                'archived_reason': 'Archived Reason',
                'archived_by': 'Archived By',
                # Factory insurance fields
                'factory_building': 'Factory - Building',
                'factory_plant_machinery': 'Factory - Plant & Machinery',
                'factory_furniture_fittings': 'Factory - Furniture & Fittings',
                'factory_stocks': 'Factory - Stocks',
                'factory_electrical_installations': 'Factory - Electrical Installations',
                # Health insurance fields
                'health_plan_type': 'Health - Plan Type',
                'health_floater_sum_insured': 'Health - Floater Sum Insured',
                'health_floater_bonus': 'Health - Floater Bonus',
                'health_floater_deductible': 'Health - Floater Deductible',
                'health_members': 'Health - Insured Members'
            }
            
            # Write headers with user-friendly names
            headers = []
            for col in df_history.columns:
                headers.append(column_mapping.get(col, col))
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Write data with date formatting
            for row_idx, (_, row) in enumerate(df_history.iterrows(), 2):
                for col_idx, (col_name, value) in enumerate(row.items(), 1):
                    # Format dates for display
                    if col_name in ['payment_date', 'policy_from', 'policy_to', 'last_reminder_sent', 'renewed_at', 'created_at', 'updated_at', 'archived_at']:
                        value = self._convert_date_for_display(value)
                    # Format boolean values
                    elif col_name in ['one_time_insurance', 'commission_received']:
                        value = 'Yes' if value else 'No'
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Created policy history sheet with {len(df_history)} records")
            
        except Exception as e:
            logger.error(f"Error creating policy history sheet: {e}")

    def check_supabase_changes(self):
        """Check if Supabase data has changed"""
        try:
            clients = self.supabase.table("clients").select("*").execute()
            members = self.supabase.table("members").select("*").execute()
            policies = self.supabase.table("policies").select("*").execute()
            pending = self.supabase.table("pending_policies").select("*").execute()
            claims = self.supabase.table("claims").select("*").execute()
            policy_history = self.supabase.table("policy_history").select("*").execute()

            current_hashes = {
                'clients': self._get_data_hash(clients.data),
                'members': self._get_data_hash(members.data),
                'policies': self._get_data_hash(policies.data),
                'pending_policies': self._get_data_hash(pending.data),
                'claims': self._get_data_hash(claims.data),
                'policy_history': self._get_data_hash(policy_history.data)
            }

            changed = current_hashes != self.last_supabase_data

            if changed:
                changed_tables = [k for k in current_hashes if current_hashes[k] != self.last_supabase_data.get(k)]
                logger.info(f"[{datetime.now().strftime('%H:%M:%S')}] Supabase changes detected in: {', '.join(changed_tables)}")
                self.export_supabase_to_excel()

        except Exception as e:
            logger.error(f"Error checking Supabase: {e}")

    def start_realtime_sync(self, poll_interval=10):
        """Start real-time synchronization in background"""
        if self.is_syncing:
            logger.warning("Sync already running")
            return

        self.is_syncing = True
        self.stop_sync = False
        
        def sync_loop():
            logger.info(f"🔄 Real-time Excel sync started (polling every {poll_interval}s)")
            while not self.stop_sync:
                try:
                    self.check_supabase_changes()
                    time.sleep(poll_interval)
                except Exception as e:
                    logger.error(f"Sync loop error: {e}")
                    time.sleep(poll_interval)
            
            logger.info("🛑 Real-time Excel sync stopped")
            self.is_syncing = False

        self.sync_thread = threading.Thread(target=sync_loop, daemon=True)
        self.sync_thread.start()

    def stop_realtime_sync(self):
        """Stop real-time synchronization"""
        self.stop_sync = True
        if self.sync_thread:
            self.sync_thread.join(timeout=5)

    def get_drive_file_info(self):
        """Get Google Drive file information (shared drive compatible)"""
        if not self.drive_service or not self.drive_file_id:
            return None
            
        try:
            file_info = self.drive_service.files().get(
                fileId=self.drive_file_id,
                fields='id,name,modifiedTime,size,webViewLink,webContentLink',
                supportsAllDrives=True  # Essential for shared drives
            ).execute()
            
            return {
                'file_id': file_info.get('id'),
                'name': file_info.get('name'),
                'modified_time': file_info.get('modifiedTime'),
                'size': file_info.get('size'),
                'view_link': file_info.get('webViewLink'),
                'download_link': file_info.get('webContentLink')
            }
            
        except Exception as e:
            logger.error(f"Error getting shared drive file info: {e}")
            return None

    def get_shareable_link(self):
        """Get shareable Google Drive link"""
        file_info = self.get_drive_file_info()
        return file_info['view_link'] if file_info else None

    def manual_sync(self):
        """Manually trigger sync"""
        logger.info("Manual sync triggered")
        self.export_supabase_to_excel()


# Global instance
excel_sync = None

def initialize_excel_sync():
    """Initialize the global Excel sync service"""
    global excel_sync
    if excel_sync is None:
        try:
            excel_sync = RealtimeExcelSync()
            excel_sync.start_realtime_sync(poll_interval=30)  # Check every 30 seconds
            logger.info("Excel sync service initialized and started")
        except Exception as e:
            logger.error(f"Failed to initialize Excel sync: {e}")
            excel_sync = None
    return excel_sync

def get_excel_sync():
    """Get the Excel sync service instance"""
    return excel_sync
